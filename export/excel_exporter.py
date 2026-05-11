"""
Excel exporter module — generates .xlsx files from analysis results.

Creates a clean Excel deliverable with:
- Column A: Reference Code (task/subtask/step ID)
- Column B: Task Title (task rows only)
- Column C: Sub Task Title (subtask rows only)
- Column D: Step Title (step rows only)
- Column E: Standards (document_filename from traceability, subtask rows only)

Each task → subtask → steps gets its own row hierarchy.
Styling: blue for tasks, green for subtasks, yellow for steps.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import io


def generate_excel(tasks: List[Dict]) -> bytes:
    """
    Generates an Excel (.xlsx) file from task hierarchy data.
    
    Args:
        tasks: List of task dicts from LLM output
    
    Returns:
        Excel file as bytes for download
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TNA Results"
    
    # Define colors
    TASK_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue
    SUBTASK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    STEP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Dark blue
    
    # Define fonts
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=11)
    
    # Define borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    
    # Add headers
    headers = [
        "Reference Code",
        "Task Title",
        "Sub Task Title",
        "Step Title",
        "Standards (Desired Performance)"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    # Row counter
    current_row = 2
    
    # Add data rows
    for task in tasks:
        task_id = task.get("task_id", "")
        task_text = task.get("task", "")
        
        # Task row
        task_row = {
            "A": task_id,
            "B": task_text,
            "C": "",
            "D": "",
            "E": "",
        }
        
        for col_letter, value in task_row.items():
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = value
            cell.font = BODY_FONT
            cell.fill = TASK_FILL
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        
        current_row += 1
        
        # Subtasks
        subtasks = task.get("subtasks", [])
        for subtask in subtasks:
            subtask_id = subtask.get("subtask_id", "")
            subtask_text = subtask.get("subtask", "")
            traceability = subtask.get("traceability", {})
            doc_filename = traceability.get("document_filename", "")
            
            # Subtask row
            subtask_row = {
                "A": subtask_id,
                "B": "",
                "C": subtask_text,
                "D": "",
                "E": doc_filename,
            }
            
            for col_letter, value in subtask_row.items():
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = value
                cell.font = BODY_FONT
                cell.fill = SUBTASK_FILL
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
            
            current_row += 1
            
            # Steps
            steps = subtask.get("steps", [])
            for step in steps:
                step_id = step.get("step_id", "")
                step_text = step.get("step", "")
                
                # Step row
                step_row = {
                    "A": step_id,
                    "B": "",
                    "C": "",
                    "D": step_text,
                    "E": "",
                }
                
                for col_letter, value in step_row.items():
                    cell = ws[f"{col_letter}{current_row}"]
                    cell.value = value
                    cell.font = BODY_FONT
                    cell.fill = STEP_FILL
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = THIN_BORDER
                
                current_row += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
